# How to write  data in workbook cell And Column Name:

from  openpyxl import Workbook
wb = Workbook()
ws = wb.active
# Define columna Names
columnsName = ["Name","Age","Email"]

# Write columns Name to first row:

for col_num,col_title in enumerate(columnsName,start=1):
    ws.cell(row=1,column=col_num,value=col_title)

# Define Data:
data = [
    ("John Doe", 30, "john.doe@example.com"),
    ("Jane Smith", 25, "jane.smith@example.com"),
    ("Emily Johnson", 35, "emily.johnson@example.com")
]
for row_num,row_data in enumerate(data,start=2):
    for col_num,cell_value in enumerate(row_data,start=1):
        ws.cell(row=row_num,column=col_num,value=cell_value)
wb.save("writeRange.xlsx")
