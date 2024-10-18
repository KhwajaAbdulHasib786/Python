# from  faker import Faker
# fake_data = Faker()
# print(fake_data.name())


# How to write  data in workbook cell.

# from  faker import  Faker
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# fake_data = Faker()
# for i in range(1,11):
#     for j in range(1,4):
#         ws.cell(row=i,column=1).value = fake_data.name()
#         ws.cell(row=i,column=2).value = fake_data.email()
#         ws.cell(row=i, column=3).value = fake_data.address()
# wb.save("hasib_data.xlsx")








# from openpyxl import Workbook
#
# # Initialize the Workbook and select the active worksheet
# wb = Workbook()
# ws = wb.active
#
# # Define column names
# column_names = ["Name", "Age", "Email","PhoneNumber"]
#
# # Write column names to the first row
# for col_num, column_title in enumerate(column_names, start=1):
#     ws.cell(row=1, column=col_num, value=column_title)
#
# # Write some sample data
# data = [
#     ("John Doe", 30, "john.doe@example.com"),
#     ("Jane Smith", 25, "jane.smith@example.com"),
#     ("Emily Johnson", 35, "emily.johnson@example.com")
# ]
#
# # Write data to the worksheet starting from the second row
# for row_num, row_data in enumerate(data, start=2):
#     for col_num, cell_value in enumerate(row_data, start=1):
#         ws.cell(row=row_num, column=col_num, value=cell_value)
#
# # Save the workbook to a file
# wb.save("sample_data.xlsx")


# How to write  data in workbook cell And Column Name:

from  openpyxl import Workbook
wb = Workbook()
ws = wb.active
# Define columna Names
columnsName = ["Name","Age","Email"]

# Write columns Name to first row:

for col_num,col_title in enumerate(columnsName,start=1):
    ws.cell(row=1,column=col_num,value=col_title)

# Define Data:
data = [
    ("John Doe", 30, "john.doe@example.com"),
    ("Jane Smith", 25, "jane.smith@example.com"),
    ("Emily Johnson", 35, "emily.johnson@example.com")
]
for row_num,row_data in enumerate(data,start=2):
    for col_num,cell_value in enumerate(row_data,start=1):
        ws.cell(row=row_num,column=col_num,value=cell_value)
wb.save("writeRange.xlsx")







